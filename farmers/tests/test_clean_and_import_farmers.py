import json
import tempfile
from io import StringIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from farmers.farmer_quarter_import import parse_quarter_workbook, run_full_import
from masters.models import Crop, District, Farmer, ProblemMaster, Village
from visits.models import Visit


def _write_quarter_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1"
    ws["A1"] = "Agri Clinic"
    ws["A2"] = "Villupuram"
    ws["A3"] = "Quarter 1 Group Summary"
    ws["A12"] = "Particulars"
    ws["A13"] = "Ananthapuram"
    ws["A14"] = "Kanagaraj Ananthapuram (9688953207)"
    ws["A15"] = "Murthy Soorapattu (9047844769)"
    ws["A16"] = "Soorapattu"
    ws["A17"] = "Chandran Agaram"
    ws["A18"] = "Ravi.P -K.Kuchipalayam 9994250059"
    wb.save(path)
    wb.close()


class FarmerQuarterImportTest(TestCase):
    def setUp(self):
        User = get_user_model()
        self.employee = User.objects.create_user("field_agent", password="pass")
        from accounts.models import EmployeeProfile

        EmployeeProfile.objects.create(user=self.employee, phone="9111111111")

        self.district = District.objects.create(name="Villupuram", is_active=True)
        self.village = Village.objects.create(
            name="OldVillage", district=self.district, is_active=True
        )
        self.farmer = Farmer.objects.create(
            name="Test Farmer",
            phone="9000000001",
            district=self.district,
            village=self.village,
            is_active=True,
        )
        Visit.objects.create(
            farmer=self.farmer,
            employee=self.employee,
            farmer_name=self.farmer.name,
            farmer_phone=self.farmer.phone,
            status="submitted",
        )

    def test_parse_quarter_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "quarter1.xlsx"
            _write_quarter_workbook(path)
            farmers, invalid, villages = parse_quarter_workbook(
                path, quarter_key="quarter1"
            )
        self.assertEqual(len(farmers), 4)
        self.assertIn("Ananthapuram", villages)
        self.assertIn("Soorapattu", villages)
        phones = {f.phone for f in farmers if f.phone}
        self.assertIn("9688953207", phones)
        self.assertIn("9994250059", phones)
        no_phone = [f for f in farmers if not f.phone]
        self.assertEqual(len(no_phone), 1)
        self.assertEqual(no_phone[0].name, "Chandran Agaram")

    def test_dry_run_command_shows_audit_and_json(self):
        out = StringIO()
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "QUARTER 1GrpSum.xlsx"
            _write_quarter_workbook(path)
            call_command(
                "clean_and_import_farmers",
                "--dry-run",
                f"--quarter1={path}",
                stdout=out,
            )
        text = out.getvalue()
        self.assertIn("=== A. KEEP", text)
        self.assertIn("=== B. DELETE", text)
        self.assertIn("masters_farmer", text)
        self.assertIn("auth_user", text)
        self.assertIn("Dry-run only", text)
        self.assertIn("farmers_created", text)

    def test_confirm_import_deletes_old_and_creates_farmers(self):
        from unittest.mock import patch

        from masters.models import ProblemCategory

        crop_before = Crop.objects.count()
        problem_before = ProblemMaster.objects.count()

        Crop.objects.create(name_en="Rice", name_ta="Rice", is_active=True)
        category, _ = ProblemCategory.objects.get_or_create(
            code="pest",
            defaults={"name": "Pest", "is_active": True},
        )
        ProblemMaster.objects.create(name="Aphids", category=category, is_active=True)

        User = get_user_model()
        admin = User.objects.create_superuser("admin", "admin@test.com", "pass")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "QUARTER 1GrpSum.xlsx"
            path2 = Path(tmp) / "QUARTER 2GrpSum.xlsx"
            _write_quarter_workbook(path)
            _write_quarter_workbook(path2)
            backup_file = Path(tmp) / "backup.sql"
            backup_file.write_text("-- test backup", encoding="utf-8")
            out = StringIO()
            with patch(
                "farmers.management.commands.clean_and_import_farmers.create_db_backup",
                return_value=backup_file,
            ):
                call_command(
                    "clean_and_import_farmers",
                    "--confirm",
                    f"--quarter1={path}",
                    f"--quarter2={path2}",
                    stdout=out,
                )

        self.assertFalse(Farmer.objects.filter(name="Test Farmer").exists())
        self.assertEqual(Farmer.objects.count(), 4)
        self.assertTrue(Crop.objects.count() >= crop_before + 1)
        self.assertTrue(ProblemMaster.objects.count() >= problem_before + 1)
        self.assertTrue(User.objects.filter(username="admin").exists())
        self.assertTrue(User.objects.filter(username="field_agent").exists())
        self.assertEqual(Visit.objects.count(), 0)

        imported = Farmer.objects.filter(phone="9688953207").first()
        self.assertIsNotNone(imported)
        self.assertIn("quarter1", imported.source_quarter)
        self.assertEqual(imported.state, "Tamil Nadu")
        self.assertEqual(imported.district.name, "Villupuram")

        text = out.getvalue()
        self.assertIn("Clean and import complete.", text)
        json_start = text.rfind('{\n  "deleted_counts"')
        self.assertGreater(json_start, -1)
        summary = json.loads(text[json_start:])
        self.assertEqual(summary["farmers_created"], 4)
        self.assertGreater(summary["village_count"], 0)
