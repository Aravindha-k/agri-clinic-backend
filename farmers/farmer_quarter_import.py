"""Parse quarter group-summary Excel files and import farmers."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from masters.models import District, Farmer, Village

DEFAULT_DISTRICT = "Villupuram"
DEFAULT_STATE = "Tamil Nadu"
DATA_START_ROW = 13

HEADER_KEYWORDS = (
    "particulars",
    "group summary",
    "group sum",
    "quarter",
    "clinic",
    "address",
    "date",
    "tamil nadu",
    "villupuram",
    "grand total",
    "total",
    "s.no",
    "sl.no",
    "serial",
    "phone",
    "mobile",
    "name",
    "village",
    "district",
)

PHONE_RE = re.compile(r"(?<!\d)([6-9]\d{9})(?!\d)")


@dataclass
class ParsedFarmerRow:
    name: str
    phone: str
    village: str
    row_num: int
    sheet: str
    raw: str


@dataclass
class ImportSummary:
    deleted_counts: dict[str, int] = field(default_factory=dict)
    quarter1_rows_processed: int = 0
    quarter2_rows_processed: int = 0
    farmers_created: int = 0
    farmers_updated: int = 0
    duplicates_skipped: int = 0
    invalid_rows: list[dict] = field(default_factory=list)
    village_count: int = 0
    villages_created: int = 0
    district: str = DEFAULT_DISTRICT

    def to_dict(self) -> dict:
        return {
            "deleted_counts": self.deleted_counts,
            "quarter1_rows_processed": self.quarter1_rows_processed,
            "quarter2_rows_processed": self.quarter2_rows_processed,
            "farmers_created": self.farmers_created,
            "farmers_updated": self.farmers_updated,
            "duplicates_skipped": self.duplicates_skipped,
            "invalid_rows": self.invalid_rows,
            "village_count": self.village_count,
            "villages_created": self.villages_created,
            "district": self.district,
        }


@dataclass
class MergeImportSummary:
    farmers_before: int = 0
    quarter3_rows_processed: int = 0
    quarter4_rows_processed: int = 0
    farmers_created: int = 0
    farmers_updated: int = 0
    duplicates_skipped: int = 0
    invalid_rows: list[dict] = field(default_factory=list)
    village_count: int = 0
    villages_created: int = 0
    district: str = DEFAULT_DISTRICT

    def to_dict(self) -> dict:
        return {
            "farmers_before": self.farmers_before,
            "quarter3_rows_processed": self.quarter3_rows_processed,
            "quarter4_rows_processed": self.quarter4_rows_processed,
            "farmers_created": self.farmers_created,
            "farmers_updated": self.farmers_updated,
            "duplicates_skipped": self.duplicates_skipped,
            "invalid_rows": self.invalid_rows,
            "village_count": self.village_count,
            "villages_created": self.villages_created,
            "district": self.district,
        }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_text(cells: tuple[Any, ...]) -> str:
    parts = [_cell_text(c) for c in cells if _cell_text(c)]
    return " ".join(parts).strip()


def _is_header_row(text: str) -> bool:
    if not text:
        return True
    low = text.lower()
    if any(kw in low for kw in HEADER_KEYWORDS):
        return True
    if re.fullmatch(r"[\d.,\s]+", text):
        return True
    return False


def _extract_phone(text: str) -> str:
    matches = PHONE_RE.findall(text)
    return matches[-1] if matches else ""


def _clean_name(text: str, phone: str) -> str:
    name = text
    if phone:
        name = name.replace(phone, "")
    name = re.sub(r"\(\s*\d*\s*\)", "", name)
    name = re.sub(r"\([^)]*\)", "", name)
    name = re.sub(r"\s+", " ", name).strip(" -.,;:")
    return name


def _classify_row(text: str, next_text: str) -> str:
    """Return skip | village | farmer."""
    if not text or _is_header_row(text):
        return "skip"
    if _extract_phone(text):
        return "farmer"
    words = text.split()
    if next_text and _extract_phone(next_text):
        if len(words) == 1:
            return "village"
        return "farmer"
    if len(words) == 1 and len(text) <= 40:
        return "village"
    return "farmer"


def parse_quarter_workbook(
    source: str | Path | BinaryIO,
    *,
    quarter_key: str,
    source_file: str = "",
) -> tuple[list[ParsedFarmerRow], list[dict], set[str]]:
    """Parse one quarter Excel file. Does not write to DB."""
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")
        wb = load_workbook(path, read_only=True, data_only=True)
        source_file = source_file or path.name
    else:
        wb = load_workbook(source, read_only=True, data_only=True)

    farmers: list[ParsedFarmerRow] = []
    invalid_rows: list[dict] = []
    villages: set[str] = set()

    for sheet in wb.worksheets:
        raw_rows: list[tuple[int, str]] = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx < DATA_START_ROW:
                continue
            text = _row_text(row)
            if text:
                raw_rows.append((row_idx, text))

        for i, (row_num, text) in enumerate(raw_rows):
            next_text = raw_rows[i + 1][1] if i + 1 < len(raw_rows) else ""
            kind = _classify_row(text, next_text)
            if kind == "skip":
                continue
            if kind == "village":
                village_name = _clean_name(text, "")
                if village_name:
                    villages.add(village_name)
                continue

            phone = _extract_phone(text)
            name = _clean_name(text, phone)
            if not name:
                invalid_rows.append(
                    {
                        "quarter": quarter_key,
                        "sheet": sheet.title,
                        "row": row_num,
                        "raw": text,
                        "reason": "Could not extract farmer name",
                    }
                )
                continue

            village_name = ""
            for j in range(i - 1, -1, -1):
                prev_num, prev_text = raw_rows[j]
                prev_next = raw_rows[j + 1][1] if j + 1 < len(raw_rows) else ""
                if _classify_row(prev_text, prev_next) == "village":
                    village_name = _clean_name(prev_text, "")
                    break

            if not village_name:
                invalid_rows.append(
                    {
                        "quarter": quarter_key,
                        "sheet": sheet.title,
                        "row": row_num,
                        "raw": text,
                        "reason": "No village heading found above farmer row",
                    }
                )
                continue

            villages.add(village_name)
            farmers.append(
                ParsedFarmerRow(
                    name=name,
                    phone=phone,
                    village=village_name,
                    row_num=row_num,
                    sheet=sheet.title,
                    raw=text,
                )
            )

    wb.close()
    return farmers, invalid_rows, villages


def _get_or_create_district(name: str = DEFAULT_DISTRICT) -> District:
    district = District.objects.filter(name__iexact=name).first()
    if district:
        return district
    return District.objects.create(name=name, is_active=True)


def _get_or_create_village(name: str, district: District) -> tuple[Village, bool]:
    village = Village.objects.filter(
        name__iexact=name, district=district
    ).first()
    if village:
        return village, False
    return Village.objects.create(name=name, district=district, is_active=True), True


def build_existing_farmer_cache() -> dict:
    """Preload existing farmers for merge-import deduplication."""
    by_phone: dict[str, Farmer] = {}
    by_name_village: dict[tuple[str, str], Farmer] = {}
    for farmer in Farmer.objects.select_related("village").iterator():
        if farmer.phone:
            by_phone[farmer.phone] = farmer
        if farmer.village_id:
            by_name_village[
                (farmer.name.lower(), farmer.village.name.lower())
            ] = farmer
    return {"by_phone": by_phone, "by_name_village": by_name_village}


def _existing_village_names(district: District | None = None) -> set[str]:
    district = district or District.objects.filter(name__iexact=DEFAULT_DISTRICT).first()
    if district is None:
        return set()
    return {
        name.lower()
        for name in Village.objects.filter(district=district).values_list(
            "name", flat=True
        )
    }


def count_new_villages(village_names: set[str], district: District | None = None) -> int:
    existing = _existing_village_names(district)
    return sum(1 for name in village_names if name.lower() not in existing)


def _merge_source(existing: Farmer, quarter_key: str, source_file: str) -> bool:
    changed = False
    quarters = {q.strip() for q in existing.source_quarter.split(",") if q.strip()}
    if quarter_key not in quarters:
        quarters.add(quarter_key)
        existing.source_quarter = ",".join(sorted(quarters))
        changed = True
    files = {f.strip() for f in existing.source_file.split(";") if f.strip()}
    if source_file and source_file not in files:
        files.add(source_file)
        existing.source_file = "; ".join(sorted(files))
        changed = True
    if not existing.state:
        existing.state = DEFAULT_STATE
        changed = True
    return changed


def _find_existing_farmer(
    row: ParsedFarmerRow,
    by_phone: dict[str, Farmer],
    by_name_village: dict[tuple[str, str], Farmer],
    *,
    query_db: bool,
) -> Farmer | None:
    existing: Farmer | None = None
    if row.phone:
        existing = by_phone.get(row.phone)
        if existing is None and query_db:
            existing = Farmer.objects.filter(phone=row.phone).first()
    if existing is None:
        key = (row.name.lower(), row.village.lower())
        existing = by_name_village.get(key)
        if existing is None and query_db:
            existing = (
                Farmer.objects.filter(name__iexact=row.name)
                .filter(village__name__iexact=row.village)
                .first()
            )
    return existing


def import_parsed_farmers(
    rows: list[ParsedFarmerRow],
    *,
    quarter_key: str,
    source_file: str,
    district: District | None = None,
    dry_run: bool = False,
    session_cache: dict | None = None,
) -> tuple[int, int, int, int]:
    """
    Import parsed rows. Returns (created, updated, duplicates_skipped, villages_created).
    session_cache tracks phone/name+village within a single import run.
    """
    district = district or _get_or_create_district()
    cache = session_cache if session_cache is not None else {}
    by_phone: dict[str, Farmer] = cache.setdefault("by_phone", {})
    by_name_village: dict[tuple[str, str], Farmer] = cache.setdefault(
        "by_name_village", {}
    )

    created = updated = skipped = villages_created = 0

    for row in rows:
        village = None
        if not dry_run:
            village, village_created = _get_or_create_village(row.village, district)
            if village_created:
                villages_created += 1

        existing = _find_existing_farmer(
            row, by_phone, by_name_village, query_db=not dry_run
        )

        if existing is not None:
            if dry_run:
                skipped += 1
                continue
            changed = _merge_source(existing, quarter_key, source_file)
            if row.phone and not existing.phone:
                existing.phone = row.phone
                changed = True
            if existing.village_id is None and village is not None:
                existing.village = village
                changed = True
            if existing.district_id is None:
                existing.district = district
                changed = True
            if changed:
                existing.save()
                updated += 1
            else:
                skipped += 1
            if row.phone:
                by_phone[row.phone] = existing
            by_name_village[(row.name.lower(), row.village.lower())] = existing
            continue

        if dry_run:
            created += 1
            continue

        farmer = Farmer.objects.create(
            name=row.name,
            phone=row.phone,
            district=district,
            village=village,
            state=DEFAULT_STATE,
            source_file=source_file,
            source_quarter=quarter_key,
            is_active=True,
        )
        created += 1
        if row.phone:
            by_phone[row.phone] = farmer
        by_name_village[(row.name.lower(), row.village.lower())] = farmer

    return created, updated, skipped, villages_created


def preview_quarter_file(path: str | Path, quarter_key: str) -> dict:
    farmers, invalid, villages = parse_quarter_workbook(
        path, quarter_key=quarter_key
    )
    return {
        "quarter": quarter_key,
        "file": str(path),
        "rows_processed": len(farmers),
        "village_count": len(villages),
        "invalid_rows": invalid,
        "sample_farmers": [
            {
                "name": f.name,
                "phone": f.phone or None,
                "village": f.village,
                "row": f.row_num,
            }
            for f in farmers[:10]
        ],
    }


def run_full_import(
    quarter1_path: str | Path | None,
    quarter2_path: str | Path | None,
    *,
    dry_run: bool = False,
) -> ImportSummary:
    summary = ImportSummary()
    district = _get_or_create_district() if not dry_run else None
    if dry_run:
        district_name = DEFAULT_DISTRICT
    else:
        district_name = district.name if district else DEFAULT_DISTRICT
    summary.district = district_name

    session_cache: dict = {}
    all_villages: set[str] = set()

    file_specs = [
        (quarter1_path, "quarter1"),
        (quarter2_path, "quarter2"),
    ]
    for path, quarter_key in file_specs:
        if not path:
            continue
        path = Path(path)
        farmers, invalid, villages = parse_quarter_workbook(
            path, quarter_key=quarter_key, source_file=path.name
        )
        all_villages |= villages
        summary.invalid_rows.extend(invalid)
        if quarter_key == "quarter1":
            summary.quarter1_rows_processed = len(farmers)
        else:
            summary.quarter2_rows_processed = len(farmers)

        created, updated, skipped, villages_created = import_parsed_farmers(
            farmers,
            quarter_key=quarter_key,
            source_file=path.name,
            district=district,
            dry_run=dry_run,
            session_cache=session_cache,
        )
        summary.farmers_created += created
        summary.farmers_updated += updated
        summary.duplicates_skipped += skipped
        summary.villages_created += villages_created

    summary.village_count = len(all_villages)
    return summary


def preview_merge_import(
    quarter3_path: str | Path | None,
    quarter4_path: str | Path | None,
) -> dict[str, Any]:
    """Dry-run merge preview for Q3/Q4 without writing to DB."""
    district = District.objects.filter(name__iexact=DEFAULT_DISTRICT).first()
    all_villages: set[str] = set()
    invalid_rows: list[dict] = []
    q3_parsed = q4_parsed = 0

    session_cache = build_existing_farmer_cache()
    created = updated = skipped = 0

    file_specs = [
        (quarter3_path, "quarter3"),
        (quarter4_path, "quarter4"),
    ]
    for path, quarter_key in file_specs:
        if not path:
            continue
        path = Path(path)
        farmers, invalid, villages = parse_quarter_workbook(
            path, quarter_key=quarter_key, source_file=path.name
        )
        all_villages |= villages
        invalid_rows.extend(invalid)
        if quarter_key == "quarter3":
            q3_parsed = len(farmers)
        else:
            q4_parsed = len(farmers)

        c, u, s, _ = import_parsed_farmers(
            farmers,
            quarter_key=quarter_key,
            source_file=path.name,
            district=district,
            dry_run=True,
            session_cache=session_cache,
        )
        created += c
        updated += u
        skipped += s

    return {
        "ready": bool(q3_parsed or q4_parsed),
        "quarter3_parsed": q3_parsed,
        "quarter4_parsed": q4_parsed,
        "farmers_created": created,
        "farmers_updated": updated,
        "duplicates_skipped": skipped,
        "invalid_rows": invalid_rows,
        "village_count": len(all_villages),
        "villages_created": count_new_villages(all_villages, district),
        "district": district.name if district else DEFAULT_DISTRICT,
    }


def run_merge_import(
    quarter3_path: str | Path | None,
    quarter4_path: str | Path | None,
    *,
    dry_run: bool = False,
) -> MergeImportSummary:
    """Merge Q3/Q4 farmers into existing live data without deleting."""
    summary = MergeImportSummary()
    summary.farmers_before = Farmer.objects.count()
    district = _get_or_create_district() if not dry_run else None
    if dry_run:
        preview = preview_merge_import(quarter3_path, quarter4_path)
        summary.quarter3_rows_processed = preview["quarter3_parsed"]
        summary.quarter4_rows_processed = preview["quarter4_parsed"]
        summary.farmers_created = preview["farmers_created"]
        summary.farmers_updated = preview["farmers_updated"]
        summary.duplicates_skipped = preview["duplicates_skipped"]
        summary.invalid_rows = preview["invalid_rows"]
        summary.village_count = preview["village_count"]
        summary.villages_created = preview["villages_created"]
        summary.district = preview["district"]
        return summary

    district = district or _get_or_create_district()
    summary.district = district.name
    session_cache = build_existing_farmer_cache()
    all_villages: set[str] = set()

    file_specs = [
        (quarter3_path, "quarter3"),
        (quarter4_path, "quarter4"),
    ]
    for path, quarter_key in file_specs:
        if not path:
            continue
        path = Path(path)
        farmers, invalid, villages = parse_quarter_workbook(
            path, quarter_key=quarter_key, source_file=path.name
        )
        all_villages |= villages
        summary.invalid_rows.extend(invalid)
        if quarter_key == "quarter3":
            summary.quarter3_rows_processed = len(farmers)
        else:
            summary.quarter4_rows_processed = len(farmers)

        created, updated, skipped, villages_created = import_parsed_farmers(
            farmers,
            quarter_key=quarter_key,
            source_file=path.name,
            district=district,
            dry_run=False,
            session_cache=session_cache,
        )
        summary.farmers_created += created
        summary.farmers_updated += updated
        summary.duplicates_skipped += skipped
        summary.villages_created += villages_created

    summary.village_count = len(all_villages)
    return summary
