"""Excel import for Problem Items (stored as ProblemMaster rows)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, BinaryIO, Optional

from openpyxl import load_workbook

from masters.models import Crop, ProblemMaster
from masters.problem_item_utils import (
    API_CATEGORY_DISEASE,
    API_CATEGORY_NUTRIENT,
    API_CATEGORY_PEST,
    get_category_for_api_code,
)

WARNING_DISEASE_NOT_IN_FILE = "Disease data not found in uploaded file"
WARNING_NUTRIENT_NOT_IN_FILE = "Nutrient issue data not found in uploaded file"

IMPORT_SCHEMAS = (
    {
        "api_category": API_CATEGORY_PEST,
        "columns": ("Crop", "English Pest Name", "Tamil Pest Name"),
        "name_col": "English Pest Name",
        "tamil_col": "Tamil Pest Name",
        "missing_warning": WARNING_DISEASE_NOT_IN_FILE,
        "pair_warning": WARNING_NUTRIENT_NOT_IN_FILE,
    },
    {
        "api_category": API_CATEGORY_DISEASE,
        "columns": ("Crop", "English Disease Name", "Tamil Disease Name"),
        "name_col": "English Disease Name",
        "tamil_col": "Tamil Disease Name",
        "missing_warning": WARNING_DISEASE_NOT_IN_FILE,
        "pair_warning": None,
    },
    {
        "api_category": API_CATEGORY_NUTRIENT,
        "columns": ("Crop", "English Nutrient Issue Name", "Tamil Nutrient Issue Name"),
        "name_col": "English Nutrient Issue Name",
        "tamil_col": "Tamil Nutrient Issue Name",
        "missing_warning": WARNING_NUTRIENT_NOT_IN_FILE,
        "pair_warning": None,
    },
)


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_or_create_crop(crop_name: str) -> tuple[Crop, bool]:
    crop_name = crop_name.strip()
    crop = Crop.objects.filter(name_en__iexact=crop_name).first()
    if crop is None:
        crop = Crop.objects.filter(name_ta__iexact=crop_name).first()
    if crop is not None:
        return crop, False
    crop = Crop.objects.create(name_en=crop_name, name_ta=crop_name)
    return crop, True


def _find_existing(category_id: int, crop_id: int | None, name: str) -> ProblemMaster | None:
    return (
        ProblemMaster.objects.filter(
            category_id=category_id,
            crop_id=crop_id,
            name__iexact=name,
        )
        .order_by("id")
        .first()
    )


@dataclass
class ImportSummary:
    total_rows: int = 0
    imported_count: int = 0
    updated_count: int = 0
    skipped_duplicates: int = 0
    failed_count: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        warning = self.warnings[0] if self.warnings else ""
        return {
            "total_rows": self.total_rows,
            "imported_count": self.imported_count,
            "updated_count": self.updated_count,
            "skipped_duplicates": self.skipped_duplicates,
            "failed_count": self.failed_count,
            "errors": self.errors,
            "warning": warning,
            "warnings": self.warnings,
        }


def _detect_schemas(header_map: dict[str, int]) -> list[dict]:
    matched = []
    for schema in IMPORT_SCHEMAS:
        if all(col in header_map for col in schema["columns"]):
            matched.append(schema)
    return matched


def import_problem_items_from_excel(file_obj: BinaryIO) -> ImportSummary:
    summary = ImportSummary()
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        summary.failed_count = 1
        summary.errors.append({"row": 0, "reason": f"Could not read Excel file: {exc}"})
        return summary

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        summary.failed_count = 1
        summary.errors.append({"row": 0, "reason": "Excel file is empty."})
        return summary

    header_map = {
        _normalize_header(cell): idx
        for idx, cell in enumerate(header_row)
        if _normalize_header(cell)
    }
    schemas = _detect_schemas(header_map)
    if not schemas:
        summary.failed_count = 1
        summary.errors.append(
            {
                "row": 1,
                "reason": (
                    "Missing required columns. Expected pest columns "
                    "(Crop, English Pest Name, Tamil Pest Name) and/or disease "
                    "or nutrient issue column sets."
                ),
            }
        )
        return summary

    detected_categories = {s["api_category"] for s in schemas}
    if API_CATEGORY_PEST in detected_categories and API_CATEGORY_DISEASE not in detected_categories:
        summary.warnings.append(WARNING_DISEASE_NOT_IN_FILE)
    if (
        API_CATEGORY_PEST in detected_categories
        and API_CATEGORY_NUTRIENT not in detected_categories
    ):
        summary.warnings.append(WARNING_NUTRIENT_NOT_IN_FILE)

    excel_row_number = 1
    for row in rows_iter:
        excel_row_number += 1
        if not row or all(_cell_text(cell) == "" for cell in row):
            continue

        for schema in schemas:
            crop_idx = header_map[schema["columns"][0]]
            name_idx = header_map[schema["name_col"]]
            tamil_idx = header_map[schema["tamil_col"]]

            crop_name = _cell_text(row[crop_idx] if crop_idx < len(row) else "")
            english_name = _cell_text(row[name_idx] if name_idx < len(row) else "")
            tamil_name = _cell_text(row[tamil_idx] if tamil_idx < len(row) else "")

            if not crop_name and not english_name and not tamil_name:
                continue

            summary.total_rows += 1

            if not crop_name:
                summary.failed_count += 1
                summary.errors.append(
                    {
                        "row": excel_row_number,
                        "reason": f"Crop is required ({schema['api_category']}).",
                    }
                )
                continue
            if not english_name:
                summary.failed_count += 1
                summary.errors.append(
                    {
                        "row": excel_row_number,
                        "reason": f"{schema['name_col']} is required.",
                    }
                )
                continue

            try:
                category = get_category_for_api_code(schema["api_category"])
            except ValueError as exc:
                summary.failed_count += 1
                summary.errors.append({"row": excel_row_number, "reason": str(exc)})
                continue

            try:
                crop, _created = _get_or_create_crop(crop_name)
            except Exception as exc:
                summary.failed_count += 1
                summary.errors.append(
                    {"row": excel_row_number, "reason": f"Could not create crop: {exc}"}
                )
                continue

            existing = _find_existing(category.id, crop.id, english_name)
            if existing:
                if tamil_name and not (existing.tamil_name or "").strip():
                    existing.tamil_name = tamil_name
                    existing.is_active = True
                    existing.save(update_fields=["tamil_name", "is_active", "updated_at"])
                    summary.updated_count += 1
                else:
                    summary.skipped_duplicates += 1
                continue

            ProblemMaster.objects.create(
                category=category,
                crop=crop,
                name=english_name,
                tamil_name=tamil_name,
                is_active=True,
            )
            summary.imported_count += 1

    return summary
